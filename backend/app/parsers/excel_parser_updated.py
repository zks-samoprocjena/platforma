"""
OBSOLETE - DO NOT USE
Excel parser for ZKS/NIS2 compliance questionnaire - Updated version.
This parser creates duplicate controls per submeasure.
Use excel_parser_v2.py instead which properly handles unique controls with M:N relationships.
Marked obsolete on 2025-07-11.
"""
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Set
from dataclasses import dataclass, asdict, field
from enum import Enum

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class SecurityLevel(str, Enum):
    """Security levels defined in the questionnaire."""

    OSNOVNA = "osnovna"
    SREDNJA = "srednja"
    NAPREDNA = "napredna"


# Mapping for mandatory values in Excel
MANDATORY_MAPPING = {
    "OBVEZNO": True,
    "DOBROVOLJNO": False,
    "OBVEZUJUĆE POD UVJETOM": True,  # Treat as mandatory
}


@dataclass
class ControlRequirement:
    """Control requirement for a specific security level."""

    security_level: SecurityLevel
    is_mandatory: bool
    is_applicable: bool = True


@dataclass
class Control:
    """Individual control with code and description."""

    code: str
    title: str
    description: str
    requirements: List[ControlRequirement] = field(default_factory=list)
    order_index: int = 0


@dataclass
class Submeasure:
    """Submeasure containing multiple controls."""

    code: str
    title: str
    description: str  # Full description from column D
    order_index: float
    controls: List[Control] = field(default_factory=list)


@dataclass
class Measure:
    """Main measure category."""

    code: str
    title: str
    description: str
    order_index: int
    submeasures: List[Submeasure] = field(default_factory=list)


@dataclass
class QuestionnaireData:
    """Complete questionnaire data structure."""

    version: str
    measures: List[Measure]

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return asdict(self)


class ExcelParserError(Exception):
    """Exception raised when Excel parsing fails."""

    pass


class ExcelParser:
    """Parser for ZKS/NIS2 compliance questionnaire Excel file."""

    def __init__(self, file_path: str | Path):
        """Initialize parser with Excel file path."""
        self.file_path = Path(file_path)
        self.workbook: Optional[openpyxl.Workbook] = None
        self.sheets: Dict[str, Worksheet] = {}
        # Track unique controls across all sheets
        self.unique_controls: Dict[str, Control] = {}

    def __enter__(self):
        """Context manager entry."""
        self.load_workbook()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        if self.workbook:
            self.workbook.close()

    def load_workbook(self) -> None:
        """Load Excel workbook and detect sheets."""
        try:
            logger.info(f"Loading Excel file: {self.file_path}")
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.detect_sheets()
            logger.info(
                f"Successfully loaded workbook with {len(self.sheets)} relevant sheets"
            )
        except Exception as e:
            raise ExcelParserError(f"Failed to load Excel file: {e}") from e

    def detect_sheets(self) -> None:
        """Detect and categorize relevant sheets."""
        if not self.workbook:
            raise ExcelParserError("Workbook not loaded")

        logger.info(f"Available sheets: {self.workbook.sheetnames}")

        # Only process security level sheets
        security_sheets = ['OSNOVNA', 'SREDNJA', 'NAPREDNA']
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in security_sheets:
                self.sheets[sheet_name] = self.workbook[sheet_name]
                logger.info(f"Found security level sheet: {sheet_name}")

        if not self.sheets:
            raise ExcelParserError("No security level sheets found in workbook")

    def parse_questionnaire(self) -> QuestionnaireData:
        """Parse the complete questionnaire from Excel."""
        logger.info("Starting questionnaire parsing")

        # Will store all measures, submeasures by their codes
        all_measures: Dict[str, Measure] = {}
        all_submeasures: Dict[str, Submeasure] = {}

        # Parse each security level sheet
        for sheet_name in ['OSNOVNA', 'SREDNJA', 'NAPREDNA']:
            if sheet_name not in self.sheets:
                logger.warning(f"Sheet {sheet_name} not found")
                continue

            level = SecurityLevel[sheet_name]
            logger.info(f"Parsing {sheet_name} sheet")

            self._parse_sheet(
                self.sheets[sheet_name], 
                level, 
                all_measures, 
                all_submeasures
            )

        # Convert to sorted lists
        measures_list = sorted(all_measures.values(), key=lambda m: m.order_index)
        
        # Sort submeasures and controls within each measure
        for measure in measures_list:
            measure.submeasures.sort(key=lambda s: s.order_index)
            for submeasure in measure.submeasures:
                submeasure.controls.sort(key=lambda c: c.order_index)

        # Log statistics
        total_submeasures = sum(len(m.submeasures) for m in measures_list)
        total_unique_controls = len(self.unique_controls)
        total_requirements = sum(len(c.requirements) for c in self.unique_controls.values())
        
        logger.info(f"Parsing complete:")
        logger.info(f"  - Measures: {len(measures_list)}")
        logger.info(f"  - Submeasures: {total_submeasures}")
        logger.info(f"  - Unique controls: {total_unique_controls}")
        logger.info(f"  - Total requirements: {total_requirements}")

        return QuestionnaireData(version="1.0", measures=measures_list)

    def _parse_sheet(
        self, 
        sheet: Worksheet, 
        level: SecurityLevel,
        all_measures: Dict[str, Measure],
        all_submeasures: Dict[str, Submeasure]
    ) -> None:
        """Parse a single sheet and merge data into collections."""
        
        current_measure = None
        current_submeasure = None
        control_order = 0

        # Start from row 2 (skip header)
        for row_num in range(2, sheet.max_row + 1):
            row_data = self._get_row_data(sheet, row_num)

            # Skip completely empty rows
            if not any(cell for cell in row_data if cell and cell.strip()):
                continue

            # Extract data from correct columns
            # Column A (index 0): Measure number (e.g., "1.0")
            measure_num = row_data[0].strip() if row_data[0] else ""
            # Column B (index 1): MJERA (Measure name)
            measure_name = row_data[1].strip() if row_data[1] else ""
            # Column C (index 2): # (submeasure number)
            submeasure_num = row_data[2].strip() if row_data[2] else ""
            # Column D (index 3): PODSKUPOVI MJERE (submeasure description)
            submeasure_desc = row_data[3].strip() if row_data[3] else ""
            # Column E (index 4): OBVEZNOST
            obligatory = row_data[4].strip() if row_data[4] else ""
            # Column F (index 5): PODSKUP MJERE SE OCJENJUJE
            evaluated = row_data[5].strip() if row_data[5] else ""
            # Column G (index 6): KONTROLE
            control_desc = row_data[6].strip() if row_data[6] else ""

            # Process new measure (check if it's a measure number like "1.0", "2.0", etc.)
            if measure_num and measure_name:
                try:
                    # Try to parse as float to verify it's a number
                    measure_num_float = float(measure_num.replace(',', '.'))
                    if 0 < measure_num_float < 20:  # Valid measure range
                        measure_code = str(int(measure_num_float))

                        if measure_code not in all_measures:
                            current_measure = Measure(
                                code=measure_code,
                                title=measure_name,
                                description=measure_name,
                                order_index=int(measure_num_float),
                                submeasures=[]
                            )
                            all_measures[measure_code] = current_measure
                        else:
                            current_measure = all_measures[measure_code]
                except ValueError:
                    # Not a valid measure number
                    pass

            # Process submeasure
            if submeasure_num and submeasure_desc and current_measure:
                submeasure_key = f"{current_measure.code}.{submeasure_num}"
                
                if submeasure_key not in all_submeasures:
                    # Parse order index from submeasure number
                    try:
                        order_index = float(submeasure_num.replace(',', '.'))
                    except ValueError:
                        order_index = 0.0

                    current_submeasure = Submeasure(
                        code=submeasure_num,
                        title=submeasure_desc[:100],  # First 100 chars as title
                        description=submeasure_desc,  # FULL description
                        order_index=order_index,
                        controls=[]
                    )
                    all_submeasures[submeasure_key] = current_submeasure
                    current_measure.submeasures.append(current_submeasure)
                else:
                    current_submeasure = all_submeasures[submeasure_key]
                
                # Reset control order for new submeasure
                control_order = 0

            # Process control
            if control_desc and current_submeasure:
                # Parse control code and title
                control_parts = control_desc.split(':', 1)
                if len(control_parts) == 2:
                    control_code = control_parts[0].strip()
                    control_title = control_parts[1].strip()

                    # Check if evaluated
                    is_applicable = evaluated.upper() == "DA" if evaluated else True

                    # Determine if mandatory
                    is_mandatory = MANDATORY_MAPPING.get(obligatory.upper(), False)

                    # Create control requirement for this level
                    requirement = ControlRequirement(
                        security_level=level,
                        is_mandatory=is_mandatory,
                        is_applicable=is_applicable
                    )

                    # Check if control already exists (deduplication)
                    if control_code in self.unique_controls:
                        # Add requirement to existing control
                        control = self.unique_controls[control_code]
                        # Check if this level requirement already exists
                        existing_req = next(
                            (r for r in control.requirements if r.security_level == level),
                            None
                        )
                        if not existing_req:
                            control.requirements.append(requirement)
                    else:
                        # Create new control
                        control_order += 1
                        control = Control(
                            code=control_code,
                            title=control_title,
                            description=control_title,
                            requirements=[requirement],
                            order_index=control_order
                        )
                        self.unique_controls[control_code] = control
                        
                    # Add control to submeasure if not already there
                    if control not in current_submeasure.controls:
                        current_submeasure.controls.append(control)

    def _get_row_data(self, sheet: Worksheet, row_num: int) -> List[str]:
        """Get row data as list of strings."""
        row_data = []
        # Read up to column O (15 columns)
        for col_num in range(1, 16):
            cell_value = sheet.cell(row_num, col_num).value
            row_data.append(str(cell_value) if cell_value is not None else "")
        return row_data

    def export_to_json(self, output_path: str | Path) -> None:
        """Export parsed data to JSON file for verification."""
        data = self.parse_questionnaire()

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data.to_dict(), f, ensure_ascii=False, indent=2)

        logger.info(f"Exported questionnaire data to: {output_path}")

    def get_statistics(self) -> Dict[str, Any]:
        """Get parsing statistics."""
        data = self.parse_questionnaire()
        
        # Count controls by security level
        level_counts = {
            'osnovna': {'mandatory': 0, 'voluntary': 0},
            'srednja': {'mandatory': 0, 'voluntary': 0},
            'napredna': {'mandatory': 0, 'voluntary': 0}
        }
        
        for control in self.unique_controls.values():
            for req in control.requirements:
                level_name = req.security_level.value
                if req.is_mandatory:
                    level_counts[level_name]['mandatory'] += 1
                else:
                    level_counts[level_name]['voluntary'] += 1
        
        return {
            'measures': len(data.measures),
            'submeasures': sum(len(m.submeasures) for m in data.measures),
            'unique_controls': len(self.unique_controls),
            'total_requirements': sum(len(c.requirements) for c in self.unique_controls.values()),
            'controls_by_level': level_counts,
            'submeasures_with_descriptions': sum(
                1 for m in data.measures 
                for s in m.submeasures 
                if s.description and len(s.description) > 100
            )
        }


def validate_questionnaire_data(data: QuestionnaireData) -> Dict[str, Any]:
    """Validate parsed questionnaire data and return validation report."""
    validation_report = {
        "is_valid": True,
        "errors": [],
        "warnings": [],
        "statistics": {},
    }

    # Check basic structure
    if not data.measures:
        validation_report["errors"].append("No measures found")
        validation_report["is_valid"] = False

    # Count unique controls
    unique_controls = set()
    total_requirements = 0
    
    for measure in data.measures:
        for submeasure in measure.submeasures:
            for control in submeasure.controls:
                unique_controls.add(control.code)
                total_requirements += len(control.requirements)

    # Statistics
    total_submeasures = sum(len(m.submeasures) for m in data.measures)

    validation_report["statistics"] = {
        "total_measures": len(data.measures),
        "total_submeasures": total_submeasures,
        "unique_controls": len(unique_controls),
        "total_requirements": total_requirements,
        "measures_with_submeasures": sum(1 for m in data.measures if m.submeasures),
        "submeasures_with_controls": sum(
            1 for m in data.measures for s in m.submeasures if s.controls
        ),
    }

    # Expected values based on analysis
    expected_measures = 13
    expected_submeasures = 99
    expected_unique_controls = 141

    if len(data.measures) != expected_measures:
        validation_report["warnings"].append(
            f"Expected {expected_measures} measures, found {len(data.measures)}"
        )

    if total_submeasures != expected_submeasures:
        validation_report["warnings"].append(
            f"Expected {expected_submeasures} submeasures, found {total_submeasures}"
        )

    if len(unique_controls) != expected_unique_controls:
        validation_report["warnings"].append(
            f"Expected {expected_unique_controls} unique controls, found {len(unique_controls)}"
        )

    # Check submeasure descriptions
    submeasures_with_desc = sum(
        1 for m in data.measures 
        for s in m.submeasures 
        if s.description and len(s.description) > 100
    )
    
    if submeasures_with_desc < expected_submeasures:
        validation_report["errors"].append(
            f"Missing submeasure descriptions: {submeasures_with_desc}/{expected_submeasures}"
        )
        validation_report["is_valid"] = False

    # Check for controls without requirements
    controls_without_requirements = []
    for measure in data.measures:
        for submeasure in measure.submeasures:
            for control in submeasure.controls:
                if not control.requirements:
                    controls_without_requirements.append(control.code)

    if controls_without_requirements:
        validation_report["errors"].append(
            f"Controls without requirements: {controls_without_requirements[:5]}"
        )
        validation_report["is_valid"] = False

    return validation_report


if __name__ == "__main__":
    # Test the updated parser
    import sys
    
    logging.basicConfig(level=logging.INFO)
    
    excel_file = Path("/mnt/shared/_Projects/ai/specijalisticki_rad/dokumentacija/službena_dokumentacija/Prilog A - Kalkulator samoprocjene(2).xlsx")
    
    if not excel_file.exists():
        print(f"Error: File not found at {excel_file}")
        sys.exit(1)
    
    try:
        with ExcelParser(excel_file) as parser:
            # Get statistics
            stats = parser.get_statistics()
            print("\n=== Parsing Statistics ===")
            print(json.dumps(stats, indent=2))
            
            # Parse and validate
            data = parser.parse_questionnaire()
            validation = validate_questionnaire_data(data)
            
            print("\n=== Validation Report ===")
            print(json.dumps(validation, indent=2))
            
            # Export to JSON for inspection
            output_file = Path("parsed_questionnaire.json")
            parser.export_to_json(output_file)
            print(f"\nExported to: {output_file}")
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()