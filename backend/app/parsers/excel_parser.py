"""
Excel parser v2 that properly handles control uniqueness and KRIP controls.
This parser creates unique controls and maps them to submeasures.
"""
import re
import logging
from typing import Dict, List, Set, Optional
from dataclasses import dataclass, field
from pathlib import Path
from enum import Enum
import openpyxl

logger = logging.getLogger(__name__)


class SecurityLevel(str, Enum):
    """Security levels defined in the questionnaire."""

    OSNOVNA = "osnovna"
    SREDNJA = "srednja"
    NAPREDNA = "napredna"


@dataclass
class ParsedControl:
    """Represents a unique control."""
    code: str
    title: str
    description: str


@dataclass
class ControlMapping:
    """Represents a control-submeasure relationship."""
    control_code: str
    submeasure_key: str  # format: "measure.submeasure"
    order_index: int
    level: str
    is_mandatory: bool
    is_applicable: bool


@dataclass
class ParsedData:
    """Container for all parsed data."""
    measures: Dict[str, dict] = field(default_factory=dict)
    submeasures: Dict[str, dict] = field(default_factory=dict)
    controls: Dict[str, ParsedControl] = field(default_factory=dict)  # Unique controls by code
    mappings: List[ControlMapping] = field(default_factory=list)      # All control-submeasure relationships


class ExcelParser:
    """Parser that properly handles control uniqueness including KRIP controls."""
    
    # Updated pattern to handle both XXX-NNN and XXXX-NNN formats
    CONTROL_PATTERN = re.compile(r'^([A-Z]{3,4}-\d{3})')
    
    def __init__(self, file_path: Path):
        self.file_path = Path(file_path)
        self.data = ParsedData()
        self._measure_order = 0
        self._submeasure_orders = {}
    
    def parse(self) -> ParsedData:
        """Parse Excel file with proper control deduplication."""
        logger.info(f"Starting parse of {self.file_path}")
        
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # Parse each security level
            for level in ['osnovna', 'srednja', 'napredna']:
                sheet_name = level.upper()
                if sheet_name in wb.sheetnames:
                    logger.info(f"Parsing sheet: {sheet_name}")
                    self._parse_sheet(wb[sheet_name], level)
                else:
                    logger.warning(f"Sheet {sheet_name} not found")
            
            wb.close()
            
            # Log summary
            logger.info(f"Parse complete: {len(self.data.measures)} measures, "
                       f"{len(self.data.submeasures)} submeasures, "
                       f"{len(self.data.controls)} unique controls, "
                       f"{len(self.data.mappings)} mappings")
            
            return self.data
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            raise
    
    def _parse_sheet(self, sheet, level: str):
        """Parse a single sheet."""
        current_measure = None
        current_submeasure_key = None
        row_count = 0
        control_count = 0
        
        for row_num in range(2, sheet.max_row + 1):
            try:
                # Extract row data
                measure_num = sheet.cell(row_num, 1).value
                measure_name = sheet.cell(row_num, 2).value
                submeasure_num = sheet.cell(row_num, 3).value
                submeasure_desc = sheet.cell(row_num, 4).value
                obligatory = sheet.cell(row_num, 5).value
                evaluated = sheet.cell(row_num, 6).value
                control_text = sheet.cell(row_num, 7).value
                
                # Process measure
                if measure_num and measure_name:
                    current_measure = self._process_measure(measure_num, measure_name)
                
                # Process submeasure
                if submeasure_num and submeasure_desc and current_measure:
                    current_submeasure_key = self._process_submeasure(
                        current_measure, submeasure_num, submeasure_desc
                    )
                
                # Process control
                if control_text and current_submeasure_key:
                    row_count += 1
                    if self._process_control(
                        control_text, current_submeasure_key, level,
                        obligatory, evaluated
                    ):
                        control_count += 1
                        
            except Exception as e:
                logger.error(f"Error parsing row {row_num}: {e}")
                continue
        
        logger.info(f"Sheet {level}: {row_count} control rows, {control_count} valid controls")
    
    def _process_measure(self, measure_num, measure_name) -> Optional[str]:
        """Process measure data and return measure code."""
        try:
            # Handle different number formats
            if isinstance(measure_num, (int, float)):
                measure_code = str(int(measure_num))
            else:
                # Handle string formats like "1,0" or "1.0"
                measure_code = str(int(float(str(measure_num).replace(',', '.'))))
            
            if measure_code not in self.data.measures:
                self._measure_order += 1
                self.data.measures[measure_code] = {
                    'code': measure_code,
                    'name': str(measure_name).strip(),
                    'description': str(measure_name).strip(),
                    'order_index': self._measure_order
                }
                logger.debug(f"Added measure: {measure_code}")
            
            return measure_code
            
        except Exception as e:
            logger.error(f"Error processing measure {measure_num}: {e}")
            return None
    
    def _process_submeasure(self, measure_code: str, submeasure_num, 
                           submeasure_desc) -> Optional[str]:
        """Process submeasure data and return submeasure key."""
        try:
            # Create submeasure key
            submeasure_key = f"{measure_code}.{submeasure_num}"
            
            if submeasure_key not in self.data.submeasures:
                # Track order per measure
                if measure_code not in self._submeasure_orders:
                    self._submeasure_orders[measure_code] = 0
                self._submeasure_orders[measure_code] += 1
                
                # Truncate long names
                name = str(submeasure_desc).strip()
                if len(name) > 100:
                    name = name[:97] + "..."
                
                self.data.submeasures[submeasure_key] = {
                    'measure_code': measure_code,
                    'code': str(submeasure_num),
                    'name': name,
                    'description': str(submeasure_desc).strip(),
                    'order_index': self._submeasure_orders[measure_code]
                }
                logger.debug(f"Added submeasure: {submeasure_key}")
            
            return submeasure_key
            
        except Exception as e:
            logger.error(f"Error processing submeasure {submeasure_num}: {e}")
            return None
    
    def _process_control(self, control_text, submeasure_key: str, level: str,
                        obligatory, evaluated) -> bool:
        """Process control and create mapping. Returns True if valid control."""
        if not control_text:
            return False
        
        # Extract control code
        match = self.CONTROL_PATTERN.match(str(control_text).strip())
        if not match:
            logger.debug(f"No control code found in: {control_text[:50]}")
            return False
        
        control_code = match.group(1)
        
        # Add unique control if not exists
        if control_code not in self.data.controls:
            # Extract title (text after the code)
            control_title = str(control_text).strip()
            if ':' in control_title:
                # Format: "POL-001: Title here"
                control_title = control_title.split(':', 1)[1].strip()
            elif control_code in control_title:
                # Remove code from beginning
                control_title = control_title.replace(control_code, '', 1).strip()
                # Remove leading punctuation
                control_title = control_title.lstrip(':- ')
            
            self.data.controls[control_code] = ParsedControl(
                code=control_code,
                title=control_title,
                description=control_title
            )
            logger.debug(f"Added control: {control_code}")
        
        # Calculate order index for this control in this submeasure at this level
        order_index = len([m for m in self.data.mappings 
                          if m.submeasure_key == submeasure_key 
                          and m.level == level]) + 1
        
        # Create mapping
        self.data.mappings.append(ControlMapping(
            control_code=control_code,
            submeasure_key=submeasure_key,
            order_index=order_index,
            level=level,
            is_mandatory=self._is_mandatory(obligatory),
            is_applicable=self._is_applicable(evaluated)
        ))
        
        return True
    
    def _is_mandatory(self, value) -> bool:
        """Determine if control is mandatory."""
        if not value:
            return False
        
        val = str(value).upper().strip()
        # Check for Croatian terms
        return val in ['OBVEZNO', 'OBVEZUJUĆE', 'OBVEZUJUĆE POD UVJETOM', 'DA']
    
    def _is_applicable(self, value) -> bool:
        """Determine if control is applicable (should be evaluated)."""
        if not value:
            return True  # Default to applicable
        
        val = str(value).upper().strip()
        # If explicitly marked as "NE" (no), it's not applicable
        return val != 'NE'


def main():
    """Test the parser."""
    import sys
    
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Test file
    test_file = Path("/tmp/questionnaire.xlsx")
    
    if not test_file.exists():
        logger.error(f"Test file not found: {test_file}")
        sys.exit(1)
    
    # Parse
    parser = ExcelParserV2(test_file)
    data = parser.parse()
    
    # Print summary
    print("\n=== PARSING RESULTS ===")
    print(f"Measures: {len(data.measures)}")
    print(f"Submeasures: {len(data.submeasures)}")
    print(f"Unique controls: {len(data.controls)}")
    print(f"Total mappings: {len(data.mappings)}")
    
    # Count by level
    level_counts = {}
    for mapping in data.mappings:
        level_counts[mapping.level] = level_counts.get(mapping.level, 0) + 1
    
    print("\nMappings by level:")
    for level, count in sorted(level_counts.items()):
        print(f"  {level}: {count}")
    
    # Show some controls
    print("\nFirst 10 controls:")
    for i, (code, control) in enumerate(sorted(data.controls.items())[:10]):
        print(f"  {code}: {control.title[:50]}...")
    
    # Check for KRIP controls
    krip_controls = [code for code in data.controls if code.startswith('KRIP')]
    print(f"\nKRIP controls found: {len(krip_controls)}")
    if krip_controls:
        print(f"  {', '.join(krip_controls)}")


if __name__ == "__main__":
    main()